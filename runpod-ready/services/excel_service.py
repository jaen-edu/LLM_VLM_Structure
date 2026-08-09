from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

REQUIRED_COLUMNS = ("번호", "아이디")
NAME_COLUMN = "이름"
RUNPOD_URL_COLUMN = "RunPod URL"
STATUS_COLUMN = "JupyterLab 접속 가능 상태"
GPU_TYPE_COLUMN = "GPU Type"
SSH_COLUMN = "SSH"


@dataclass(slots=True)
class StudentRecord:
    row_index: int
    number: int
    user_id: str
    name: str | None
    runpod_url: str | None
    jupyter_status: str | None
    gpu_type: str | None
    ssh_command: str | None

    @property
    def is_instructor(self) -> bool:
        return self.number == 0 and self.user_id.lower() == "instructor"


class ExcelService:
    """Load, validate, and update xlsx records for RunPod operations."""

    def __init__(self, sheet_name: str = "시트1") -> None:
        self.sheet_name = sheet_name

    def load_records(self, input_path: Path) -> tuple[Workbook, Worksheet, list[StudentRecord], dict[str, int]]:
        workbook = load_workbook(input_path)
        if self.sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {self.sheet_name}")

        worksheet = workbook[self.sheet_name]
        header_map = self._read_header_map(worksheet)
        self._ensure_name_column(worksheet, header_map)
        self._ensure_runpod_url_column(worksheet, header_map)
        self._validate_required_columns(header_map)
        self._ensure_status_column(worksheet, header_map)
        self._ensure_gpu_type_column(worksheet, header_map)
        self._ensure_ssh_column(worksheet, header_map)

        records: list[StudentRecord] = []
        for row_index in range(2, worksheet.max_row + 1):
            values = {name: worksheet.cell(row=row_index, column=header_map[name]).value for name in REQUIRED_COLUMNS}
            values[NAME_COLUMN] = worksheet.cell(row=row_index, column=header_map[NAME_COLUMN]).value
            values[RUNPOD_URL_COLUMN] = worksheet.cell(row=row_index, column=header_map[RUNPOD_URL_COLUMN]).value
            values[STATUS_COLUMN] = worksheet.cell(row=row_index, column=header_map[STATUS_COLUMN]).value
            values[GPU_TYPE_COLUMN] = worksheet.cell(row=row_index, column=header_map[GPU_TYPE_COLUMN]).value
            values[SSH_COLUMN] = worksheet.cell(row=row_index, column=header_map[SSH_COLUMN]).value
            if self._is_blank_row(values):
                continue

            number = self._parse_number(values["번호"], row_index)

            raw_user_id = self._optional_text(values["아이디"])
            raw_name = self._optional_text(values[NAME_COLUMN])

            # Some operational sheets keep only number=0 for instructor row.
            if number == 0:
                raw_user_id = raw_user_id or "instructor"
                raw_name = raw_name or "instructor"

            user_id = self._required_text(raw_user_id, "아이디", row_index)
            name = self._optional_text(raw_name)
            runpod_url = self._optional_text(values[RUNPOD_URL_COLUMN])
            jupyter_status = self._optional_text(values[STATUS_COLUMN])
            gpu_type = self._optional_text(values[GPU_TYPE_COLUMN])
            ssh_command = self._optional_text(values[SSH_COLUMN])

            records.append(
                StudentRecord(
                    row_index=row_index,
                    number=number,
                    user_id=user_id,
                    name=name,
                    runpod_url=runpod_url,
                    jupyter_status=jupyter_status,
                    gpu_type=gpu_type,
                    ssh_command=ssh_command,
                )
            )

        if not records:
            raise ValueError("No valid data rows found from row 2.")

        if not any(record.is_instructor for record in records):
            raise ValueError("Instructor row is required: number=0 and user_id=instructor")

        return workbook, worksheet, records, header_map

    def update_runpod_url(
        self,
        worksheet: Worksheet,
        header_map: dict[str, int],
        record: StudentRecord,
        runpod_url: str,
    ) -> None:
        worksheet.cell(row=record.row_index, column=header_map[RUNPOD_URL_COLUMN]).value = runpod_url

    def clear_runpod_url(
        self,
        worksheet: Worksheet,
        header_map: dict[str, int],
        record: StudentRecord,
    ) -> None:
        worksheet.cell(row=record.row_index, column=header_map[RUNPOD_URL_COLUMN]).value = ""

    def update_jupyter_status(
        self,
        worksheet: Worksheet,
        header_map: dict[str, int],
        record: StudentRecord,
        status: str,
    ) -> None:
        worksheet.cell(row=record.row_index, column=header_map[STATUS_COLUMN]).value = status

    def update_gpu_type(
        self,
        worksheet: Worksheet,
        header_map: dict[str, int],
        record: StudentRecord,
        gpu_type: str,
    ) -> None:
        cell = worksheet.cell(row=record.row_index, column=header_map[GPU_TYPE_COLUMN])
        existing = self._optional_text(cell.value)
        candidate = self._optional_text(gpu_type)
        if not candidate:
            return

        # Keep existing value when only display format differs (e.g. NVIDIA A100 80GB PCIe vs A100 PCIe).
        if existing and self._is_equivalent_gpu_label(existing, candidate):
            return

        cell.value = candidate

    def clear_gpu_type(
        self,
        worksheet: Worksheet,
        header_map: dict[str, int],
        record: StudentRecord,
    ) -> None:
        worksheet.cell(row=record.row_index, column=header_map[GPU_TYPE_COLUMN]).value = ""

    def update_ssh_command(
        self,
        worksheet: Worksheet,
        header_map: dict[str, int],
        record: StudentRecord,
        ssh_command: str,
    ) -> None:
        worksheet.cell(row=record.row_index, column=header_map[SSH_COLUMN]).value = ssh_command

    def clear_ssh_command(
        self,
        worksheet: Worksheet,
        header_map: dict[str, int],
        record: StudentRecord,
    ) -> None:
        worksheet.cell(row=record.row_index, column=header_map[SSH_COLUMN]).value = ""

    def save_updated_copy(
        self,
        workbook: Workbook,
        input_path: Path,
        output_path: Path | None = None,
    ) -> Path:
        target = output_path or self.derive_output_path(input_path)
        target.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(target)
        return target

    @staticmethod
    def derive_output_path(input_path: Path) -> Path:
        # Single-file workflow: update the input workbook in place by default.
        return input_path

    @staticmethod
    def _read_header_map(worksheet: Worksheet) -> dict[str, int]:
        header_map: dict[str, int] = {}
        for column in range(1, worksheet.max_column + 1):
            value = worksheet.cell(row=1, column=column).value
            if value is None:
                continue
            text = str(value).strip()
            if text:
                header_map[text] = column
        return header_map

    @staticmethod
    @staticmethod
    def _ensure_name_column(worksheet: Worksheet, header_map: dict[str, int]) -> None:
        if NAME_COLUMN in header_map:
            return

        user_id_col = header_map.get("아이디")
        if user_id_col is None:
            insert_col = max((col for col in header_map.values()), default=0) + 1
        else:
            insert_col = user_id_col + 1
            worksheet.insert_cols(insert_col)
            for key, col in list(header_map.items()):
                if col >= insert_col:
                    header_map[key] = col + 1

        worksheet.cell(row=1, column=insert_col).value = NAME_COLUMN
        header_map[NAME_COLUMN] = insert_col

    @staticmethod
    def _ensure_runpod_url_column(worksheet: Worksheet, header_map: dict[str, int]) -> None:
        if RUNPOD_URL_COLUMN in header_map:
            return

        name_col = header_map.get(NAME_COLUMN)
        if name_col is None:
            insert_col = max((col for col in header_map.values()), default=0) + 1
        else:
            insert_col = name_col + 1
            worksheet.insert_cols(insert_col)
            for key, col in list(header_map.items()):
                if col >= insert_col:
                    header_map[key] = col + 1

        worksheet.cell(row=1, column=insert_col).value = RUNPOD_URL_COLUMN
        header_map[RUNPOD_URL_COLUMN] = insert_col

    @staticmethod
    def _validate_required_columns(header_map: dict[str, int]) -> None:
        missing = [column for column in REQUIRED_COLUMNS if column not in header_map]
        if missing:
            joined = ", ".join(missing)
            raise ValueError(f"Missing required columns: {joined}")

    @staticmethod
    def _ensure_status_column(worksheet: Worksheet, header_map: dict[str, int]) -> None:
        if STATUS_COLUMN in header_map:
            return

        url_col = header_map[RUNPOD_URL_COLUMN]
        insert_col = url_col + 1
        worksheet.insert_cols(insert_col)
        worksheet.cell(row=1, column=insert_col).value = STATUS_COLUMN

        for key, col in list(header_map.items()):
            if col >= insert_col:
                header_map[key] = col + 1
        header_map[STATUS_COLUMN] = insert_col

    @staticmethod
    def _ensure_gpu_type_column(worksheet: Worksheet, header_map: dict[str, int]) -> None:
        if GPU_TYPE_COLUMN in header_map:
            return

        status_col = header_map[STATUS_COLUMN]
        insert_col = status_col + 1
        worksheet.insert_cols(insert_col)
        worksheet.cell(row=1, column=insert_col).value = GPU_TYPE_COLUMN

        for key, col in list(header_map.items()):
            if col >= insert_col:
                header_map[key] = col + 1
        header_map[GPU_TYPE_COLUMN] = insert_col

    @staticmethod
    def _ensure_ssh_column(worksheet: Worksheet, header_map: dict[str, int]) -> None:
        if SSH_COLUMN in header_map:
            return

        gpu_col = header_map[GPU_TYPE_COLUMN]
        insert_col = gpu_col + 1
        worksheet.insert_cols(insert_col)
        worksheet.cell(row=1, column=insert_col).value = SSH_COLUMN

        for key, col in list(header_map.items()):
            if col >= insert_col:
                header_map[key] = col + 1
        header_map[SSH_COLUMN] = insert_col

    @staticmethod
    def _is_blank_row(values: dict[str, object]) -> bool:
        return all(value is None or str(value).strip() == "" for value in values.values())

    @staticmethod
    def _optional_text(value: object) -> str | None:
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    @staticmethod
    def _gpu_tokens(value: str) -> set[str]:
        return {token for token in re.findall(r"[a-z0-9]+", value.lower()) if token}

    @staticmethod
    def _is_equivalent_gpu_label(existing: str, candidate: str) -> bool:
        a = ExcelService._gpu_tokens(existing)
        b = ExcelService._gpu_tokens(candidate)
        if not a or not b:
            return False
        return a.issubset(b) or b.issubset(a)

    @staticmethod
    def _required_text(value: object, column_name: str, row_index: int) -> str:
        text = ExcelService._optional_text(value)
        if text is None:
            raise ValueError(f"Row {row_index}: '{column_name}' is required")
        return text

    @staticmethod
    def _parse_number(value: object, row_index: int) -> int:
        if value is None or str(value).strip() == "":
            raise ValueError(f"Row {row_index}: '번호' is required")

        if isinstance(value, int):
            return value

        if isinstance(value, float):
            if value.is_integer():
                return int(value)
            raise ValueError(f"Row {row_index}: invalid 번호 value '{value}'")

        text = str(value).strip()
        if re.fullmatch(r"-?\d+(\.0+)?", text):
            return int(float(text))

        try:
            return int(text)
        except ValueError as exc:
            raise ValueError(f"Row {row_index}: invalid 번호 value '{value}'") from exc
