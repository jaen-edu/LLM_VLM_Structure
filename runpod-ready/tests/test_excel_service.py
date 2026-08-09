from pathlib import Path

import pytest
from openpyxl import Workbook

from services.excel_service import ExcelService


def _write_valid_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "시트1"
    worksheet.append(["번호", "아이디"])
    worksheet.append([1, "user01"])
    worksheet.append([None, None])
    worksheet.append([0, "instructor"])
    workbook.save(path)


def test_load_records_parses_and_skips_blank_rows(tmp_path: Path) -> None:
    file_path = tmp_path / "students.xlsx"
    _write_valid_workbook(file_path)

    service = ExcelService()
    _, _, records, header_map = service.load_records(file_path)

    assert len(records) == 2
    assert records[0].number == 1
    assert records[1].is_instructor is True
    assert "이름" in header_map
    assert "RunPod URL" in header_map
    assert "JupyterLab 접속 가능 상태" in header_map
    assert "GPU Type" in header_map
    assert "SSH" in header_map


def test_load_records_validates_required_columns(tmp_path: Path) -> None:
    # 필수 입력은 번호/아이디만 허용
    file_path = tmp_path / "invalid.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "시트1"
    worksheet.append(["번호", "이름"])  # 아이디 누락
    worksheet.append([1, "홍길동"])
    workbook.save(file_path)

    service = ExcelService()
    with pytest.raises(ValueError, match="Missing required columns"):
        service.load_records(file_path)


def test_load_records_auto_creates_runpod_url_column(tmp_path: Path) -> None:
    # "RunPod URL" 컬럼이 없어도 자동 생성됨
    file_path = tmp_path / "no_url_col.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "시트1"
    worksheet.append(["번호", "아이디"])
    worksheet.append([0, "instructor"])
    workbook.save(file_path)

    service = ExcelService()
    _, _, records, header_map = service.load_records(file_path)
    assert "이름" in header_map
    assert "RunPod URL" in header_map
    assert records[0].runpod_url is None


def test_save_updated_copy_uses_input_path_by_default(tmp_path: Path) -> None:
    file_path = tmp_path / "students.xlsx"
    _write_valid_workbook(file_path)

    service = ExcelService()
    workbook, _, records, header_map = service.load_records(file_path)
    worksheet = workbook["시트1"]

    service.update_runpod_url(worksheet, header_map, records[0], "https://example.com")
    saved = service.save_updated_copy(workbook, file_path)

    assert saved.name == "students.xlsx"
    assert saved.exists()


def test_derive_output_path_returns_input_path(tmp_path: Path) -> None:
    already_updates = tmp_path / "runpod-urls.updates.xlsx"
    result = ExcelService.derive_output_path(already_updates)
    assert result == already_updates

    plain = tmp_path / "runpod-urls.xlsx"
    result2 = ExcelService.derive_output_path(plain)
    assert result2 == plain


def test_update_jupyter_status_writes_status_column(tmp_path: Path) -> None:
    file_path = tmp_path / "students.xlsx"
    _write_valid_workbook(file_path)

    service = ExcelService()
    workbook, worksheet, records, header_map = service.load_records(file_path)

    service.update_jupyter_status(worksheet, header_map, records[0], "OK")
    saved = service.save_updated_copy(workbook, file_path)

    loaded = service.load_records(saved)[0]
    ws = loaded["시트1"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    status_col = headers.index("JupyterLab 접속 가능 상태") + 1
    assert ws.cell(row=2, column=status_col).value == "OK"


def test_update_gpu_type_writes_gpu_column(tmp_path: Path) -> None:
    file_path = tmp_path / "students.xlsx"
    _write_valid_workbook(file_path)

    service = ExcelService()
    workbook, worksheet, records, header_map = service.load_records(file_path)

    service.update_gpu_type(worksheet, header_map, records[0], "NVIDIA A100 80GB PCIe")
    saved = service.save_updated_copy(workbook, file_path)

    loaded = service.load_records(saved)[0]
    ws = loaded["시트1"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    gpu_col = headers.index("GPU Type") + 1
    assert ws.cell(row=2, column=gpu_col).value == "NVIDIA A100 80GB PCIe"


def test_update_ssh_command_writes_ssh_column(tmp_path: Path) -> None:
    file_path = tmp_path / "students.xlsx"
    _write_valid_workbook(file_path)

    service = ExcelService()
    workbook, worksheet, records, header_map = service.load_records(file_path)

    service.update_ssh_command(
        worksheet,
        header_map,
        records[0],
        "ssh t5jlk6quhxubew-64411d38@ssh.runpod.io -i ~/.ssh/id_ed25519",
    )
    saved = service.save_updated_copy(workbook, file_path)

    loaded = service.load_records(saved)[0]
    ws = loaded["시트1"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    ssh_col = headers.index("SSH") + 1
    assert ws.cell(row=2, column=ssh_col).value == "ssh t5jlk6quhxubew-64411d38@ssh.runpod.io -i ~/.ssh/id_ed25519"


def test_update_gpu_type_keeps_existing_when_equivalent_alias(tmp_path: Path) -> None:
    file_path = tmp_path / "students.xlsx"
    _write_valid_workbook(file_path)

    service = ExcelService()
    workbook, worksheet, records, header_map = service.load_records(file_path)
    service.update_gpu_type(worksheet, header_map, records[0], "NVIDIA A100 80GB PCIe")
    service.update_gpu_type(worksheet, header_map, records[0], "A100 PCIe")
    saved = service.save_updated_copy(workbook, file_path)

    loaded = service.load_records(saved)[0]
    ws = loaded["시트1"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    gpu_col = headers.index("GPU Type") + 1
    assert ws.cell(row=2, column=gpu_col).value == "NVIDIA A100 80GB PCIe"


def test_load_records_accepts_integer_like_float_number(tmp_path: Path) -> None:
    file_path = tmp_path / "float-number.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "시트1"
    worksheet.append(["번호", "아이디"])
    worksheet.append([15.0, "user15"])
    worksheet.append([0, "instructor"])
    workbook.save(file_path)

    service = ExcelService()
    _, _, records, _ = service.load_records(file_path)

    assert records[0].number == 15


def test_load_records_autofills_instructor_row_when_id_missing(tmp_path: Path) -> None:
    file_path = tmp_path / "instructor-missing-id.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "시트1"
    worksheet.append(["번호", "아이디"])
    worksheet.append([1, "user01"])
    worksheet.append([0, None])
    workbook.save(file_path)

    service = ExcelService()
    _, _, records, _ = service.load_records(file_path)

    instructor = records[1]
    assert instructor.number == 0
    assert instructor.user_id == "instructor"
    assert instructor.name == "instructor"
