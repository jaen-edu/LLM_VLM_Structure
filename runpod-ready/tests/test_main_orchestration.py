from __future__ import annotations

import io
import logging
from argparse import Namespace
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from main import run
from utils.retry import RetryError


class FakeRunPodService:
    def __init__(self, existing_pods: dict[str, dict[str, Any]] | None = None) -> None:
        self._pods = existing_pods or {}
        self.calls: list[dict[str, Any]] = []
        self.stopped_ids: list[str] = []
        self.resumed_ids: list[str] = []
        self.terminated_ids: list[str] = []
        self._pod_gpu_by_id: dict[str, str] = {}
        for pod in self._pods.values():
            pod_id = str(pod.get("id", "")).strip()
            machine = pod.get("machine") or {}
            gpu = machine.get("gpuDisplayName")
            if pod_id and gpu:
                self._pod_gpu_by_id[pod_id] = str(gpu)

    def list_gpu_types(self) -> list[str]:
        return ["NVIDIA A100 80GB PCIe"]

    def list_gpu_types_with_memory(self, *, min_memory_gb: float | None = None) -> list[tuple[str, float]]:
        items = [("NVIDIA A100 80GB PCIe", 80.0), ("NVIDIA RTX 4090", 24.0)]
        if min_memory_gb is None:
            return items
        return [item for item in items if item[1] >= min_memory_gb]

    def list_gpu_types_with_memory_and_secure_price(
        self,
        *,
        min_memory_gb: float | None = None,
    ) -> list[tuple[str, float, float | None]]:
        items = [("NVIDIA A100 80GB PCIe", 80.0, 1.39), ("NVIDIA RTX 4090", 24.0, 0.59)]
        if min_memory_gb is None:
            return items
        return [item for item in items if item[1] >= min_memory_gb]

    def get_existing_pods_map(self) -> dict[str, dict[str, Any]]:
        return dict(self._pods)

    def create_or_get_pod(
        self,
        *,
        spec: Any,
        existing_pod: dict[str, Any] | None,
        force_recreate: bool,
        dry_run: bool,
    ) -> dict[str, Any]:
        self.calls.append(
            {
                "name": spec.name,
                "gpu_type_id": spec.gpu_type_id,
                "force_recreate": force_recreate,
                "dry_run": dry_run,
            }
        )
        pod = existing_pod or {
            "id": f"generated-{spec.name}",
            "name": spec.name,
            "machine": {"gpuDisplayName": spec.gpu_type_id},
            "runtime": {
                "host": "pod.example.com",
                "port": 8888,
            },
        }
        pod_id = str(pod.get("id", "")).strip()
        if pod_id:
            machine = pod.get("machine") or {}
            gpu = machine.get("gpuDisplayName")
            if gpu:
                self._pod_gpu_by_id[pod_id] = str(gpu)
        return pod

    def wait_for_runtime(self, pod_id: str, *, timeout_seconds: int) -> dict[str, Any]:
        gpu = self._pod_gpu_by_id.get(pod_id, "NVIDIA A100 80GB PCIe")
        return {
            "id": pod_id,
            "machine": {"gpuDisplayName": gpu},
            "runtime": {
                "host": "pod.example.com",
                "port": 8888,
                "token": "token-abc",
            },
        }

    def get_pod(self, pod_id: str) -> dict[str, Any]:
        for pod in self._pods.values():
            if pod.get("id") == pod_id:
                return pod
        return {
            "id": pod_id,
            "runtime": {
                "host": "pod.example.com",
                "port": 8888,
                "token": "token-abc",
            },
        }

    def stop_pod(self, pod_id: str, *, dry_run: bool) -> None:
        self.stopped_ids.append(pod_id)

    def resume_pod(self, pod_id: str, *, dry_run: bool, gpu_count: int = 1) -> None:
        self.resumed_ids.append(pod_id)

    def start_pod(self, pod_id: str, *, dry_run: bool, gpu_count: int = 1) -> None:
        self.resumed_ids.append(pod_id)

    def terminate_pod(self, pod_id: str, *, dry_run: bool) -> None:
        self.terminated_ids.append(pod_id)


class FailingFirstCreateRunPodService(FakeRunPodService):
    def __init__(self) -> None:
        super().__init__(existing_pods={})
        self._failed = False

    def create_or_get_pod(
        self,
        *,
        spec: Any,
        existing_pod: dict[str, Any] | None,
        force_recreate: bool,
        dry_run: bool,
    ) -> dict[str, Any]:
        if not self._failed:
            self._failed = True
            raise RuntimeError("capacity unavailable")
        return super().create_or_get_pod(
            spec=spec,
            existing_pod=existing_pod,
            force_recreate=force_recreate,
            dry_run=dry_run,
        )


class FallbackAwareRunPodService(FakeRunPodService):
    def __init__(self, unavailable_gpu: str) -> None:
        super().__init__(existing_pods={})
        self.unavailable_gpu = unavailable_gpu
        self.gpu_calls: list[str] = []

    def create_or_get_pod(
        self,
        *,
        spec: Any,
        existing_pod: dict[str, Any] | None,
        force_recreate: bool,
        dry_run: bool,
    ) -> dict[str, Any]:
        self.gpu_calls.append(spec.gpu_type_id)
        if spec.gpu_type_id == self.unavailable_gpu:
            raise RuntimeError("There are no longer any instances available with the requested specifications")
        return super().create_or_get_pod(
            spec=spec,
            existing_pod=existing_pod,
            force_recreate=force_recreate,
            dry_run=dry_run,
        )


class WrappedFallbackAwareRunPodService(FakeRunPodService):
    def __init__(self, unavailable_gpu: str) -> None:
        super().__init__(existing_pods={})
        self.unavailable_gpu = unavailable_gpu
        self.gpu_calls: list[str] = []

    def create_or_get_pod(
        self,
        *,
        spec: Any,
        existing_pod: dict[str, Any] | None,
        force_recreate: bool,
        dry_run: bool,
    ) -> dict[str, Any]:
        self.gpu_calls.append(spec.gpu_type_id)
        if spec.gpu_type_id == self.unavailable_gpu:
            cause = RuntimeError("There are no longer any instances available with the requested specifications")
            raise RetryError("create_pod:class-00-instructor failed after 3 attempts") from cause
        return super().create_or_get_pod(
            spec=spec,
            existing_pod=existing_pod,
            force_recreate=force_recreate,
            dry_run=dry_run,
        )


class CapacityBlockedResumeRunPodService(FakeRunPodService):
    def __init__(self, existing_pods: dict[str, dict[str, Any]], blocked_pod_ids: set[str]) -> None:
        super().__init__(existing_pods=existing_pods)
        self.blocked_pod_ids = blocked_pod_ids

    def resume_pod(self, pod_id: str, *, dry_run: bool, gpu_count: int = 1) -> None:
        self.resumed_ids.append(pod_id)
        if pod_id in self.blocked_pod_ids:
            cause = RuntimeError("There are not enough free GPUs on the host machine to start this pod.")
            raise RetryError(f"resume_pod:{pod_id} failed after 3 attempts") from cause

    def start_pod(self, pod_id: str, *, dry_run: bool, gpu_count: int = 1) -> None:
        self.resume_pod(pod_id, dry_run=dry_run, gpu_count=gpu_count)


class CapacityBlockedResumeWithFallbackRunPodService(CapacityBlockedResumeRunPodService):
    def create_or_get_pod(
        self,
        *,
        spec: Any,
        existing_pod: dict[str, Any] | None,
        force_recreate: bool,
        dry_run: bool,
    ) -> dict[str, Any]:
        self.calls.append(
            {
                "name": spec.name,
                "gpu_type_id": spec.gpu_type_id,
                "force_recreate": force_recreate,
                "dry_run": dry_run,
            }
        )
        if force_recreate and existing_pod is not None:
            old_pod_id = str(existing_pod.get("id", "")).strip()
            if old_pod_id:
                self.terminated_ids.append(old_pod_id)
            pod = {
                "id": f"recreated-{spec.name}",
                "name": spec.name,
                "machine": {"gpuDisplayName": spec.gpu_type_id},
                "runtime": {
                    "host": "pod.example.com",
                    "port": 8888,
                },
            }
            self._pods[spec.name.lower()] = pod
            self._pod_gpu_by_id[pod["id"]] = spec.gpu_type_id
            return pod

        return super().create_or_get_pod(
            spec=spec,
            existing_pod=existing_pod,
            force_recreate=force_recreate,
            dry_run=dry_run,
        )


def _write_workbook(path: Path, with_existing_url: bool = False) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "시트1"
    ws.append(["번호", "아이디", "이름", "성별", "이메일", "RunPod URL"])
    ws.append([1, "user01", "홍길동", "M", "user01@example.com", "https://existing-url" if with_existing_url else None])
    ws.append([0, "instructor", "강사", "F", "teacher@example.com", None])
    wb.save(path)


def _set_status(path: Path, row: int, status: str) -> None:
    wb = load_workbook(path)
    ws = wb["시트1"]
    ws.cell(row=1, column=7).value = "JupyterLab 접속 가능 상태"
    ws.cell(row=row, column=7).value = status
    wb.save(path)


def _base_args(input_path: Path, output_path: Path) -> Namespace:
    return Namespace(
        input=str(input_path),
        output=str(output_path),
        action="provision",
        numbers=[],
        ids=[],
        all=False,
        overwrite_url=False,
        dry_run=False,
        recreate=[],
        recreate_ids=[],
        recreate_if_unhealthy=False,
        gpu_type="NVIDIA A100 80GB PCIe",
        gpu_type_fallback=[],
        timeout=30,
        jupyter_check=False,
        recreate_on_unreachable=False,
        list_gpu_types=False,
    )


def test_run_reprovisions_when_status_is_recovery_target(tmp_path: Path, monkeypatch: Any) -> None:
    monkeypatch.setenv("RUNPOD_API_KEY", "test-key")
    input_path = tmp_path / "students.xlsx"
    output_path = tmp_path / "students.updated.xlsx"
    _write_workbook(input_path, with_existing_url=True)
    _set_status(input_path, row=2, status="UNREACHABLE")

    fake_service = FakeRunPodService(existing_pods={
        "class-01-user01": {"id": "pod-existing", "name": "class-01-user01", "runtime": {"host": "pod.example.com", "port": 8888}}
    })

    exit_code = run(
        _base_args(input_path, output_path),
        runpod_factory=lambda config, logger: fake_service,
        jupyter_checker=lambda url, timeout: (True, "HTTP_200"),
    )

    assert exit_code == 0
    assert any(call["name"] == "class-01-user01" and call["force_recreate"] for call in fake_service.calls)


def test_run_logs_status_based_reprovision_reason(tmp_path: Path, monkeypatch: Any) -> None:
    monkeypatch.setenv("RUNPOD_API_KEY", "test-key")
    input_path = tmp_path / "students.xlsx"
    output_path = tmp_path / "students.updated.xlsx"
    _write_workbook(input_path, with_existing_url=True)
    _set_status(input_path, row=2, status="UNREACHABLE")

    fake_service = FakeRunPodService(existing_pods={
        "class-01-user01": {"id": "pod-existing", "name": "class-01-user01", "runtime": {"host": "pod.example.com", "port": 8888}}
    })
    log_stream = io.StringIO()
    logger = logging.getLogger("test-status-based-reprovision")
    logger.handlers.clear()
    logger.setLevel(logging.INFO)
    logger.propagate = False
    handler = logging.StreamHandler(log_stream)
    logger.addHandler(handler)

    try:
        exit_code = run(
            _base_args(input_path, output_path),
            runpod_factory=lambda config, logger: fake_service,
            jupyter_checker=lambda url, timeout: (True, "HTTP_200"),
            logger=logger,
        )
    finally:
        logger.removeHandler(handler)

    assert exit_code == 0
    assert "status is 'UNREACHABLE' (reprovision target)" in log_stream.getvalue()


def test_run_does_not_reprovision_when_status_is_stopped(tmp_path: Path, monkeypatch: Any) -> None:
    monkeypatch.setenv("RUNPOD_API_KEY", "test-key")
    input_path = tmp_path / "students.xlsx"
    output_path = tmp_path / "students.updated.xlsx"
    _write_workbook(input_path, with_existing_url=True)
    _set_status(input_path, row=2, status="STOPPED")

    fake_service = FakeRunPodService(existing_pods={
        "class-01-user01": {"id": "pod-existing", "name": "class-01-user01", "runtime": {"host": "pod.example.com", "port": 8888}}
    })

    exit_code = run(
        _base_args(input_path, output_path),
        runpod_factory=lambda config, logger: fake_service,
        jupyter_checker=lambda url, timeout: (True, "HTTP_200"),
    )

    assert exit_code == 0
    assert all(call["name"] != "class-01-user01" for call in fake_service.calls)


def test_run_skips_existing_url_when_status_is_ok(tmp_path: Path, monkeypatch: Any) -> None:
    monkeypatch.setenv("RUNPOD_API_KEY", "test-key")
    input_path = tmp_path / "students.xlsx"
    output_path = tmp_path / "students.updated.xlsx"
    _write_workbook(input_path, with_existing_url=True)
    _set_status(input_path, row=2, status="OK")

    fake_service = FakeRunPodService(existing_pods={
        "class-01-user01": {"id": "pod-existing", "name": "class-01-user01", "runtime": {"host": "pod.example.com", "port": 8888}}
    })

    exit_code = run(
        _base_args(input_path, output_path),
        runpod_factory=lambda config, logger: fake_service,
        jupyter_checker=lambda url, timeout: (True, "HTTP_200"),
    )

    assert exit_code == 0
    assert any(call["name"] == "class-00-instructor" for call in fake_service.calls)
    assert all(call["name"] != "class-01-user01" for call in fake_service.calls)


def test_run_recreates_selected_number(tmp_path: Path, monkeypatch: Any) -> None:
    monkeypatch.setenv("RUNPOD_API_KEY", "test-key")
    input_path = tmp_path / "students.xlsx"
    output_path = tmp_path / "students.updated.xlsx"
    _write_workbook(input_path, with_existing_url=False)

    args = _base_args(input_path, output_path)
    args.recreate = [1]

    fake_service = FakeRunPodService(existing_pods={
        "class-01-user01": {"id": "pod-existing", "name": "class-01-user01", "status": "running"}
    })

    exit_code = run(
        args,
        runpod_factory=lambda config, logger: fake_service,
        jupyter_checker=lambda url, timeout: (True, "HTTP_200"),
    )

    assert exit_code == 0
    assert any(call["name"] == "class-01-user01" and call["force_recreate"] for call in fake_service.calls)

    saved = load_workbook(output_path)
    ws = saved["시트1"]
    assert ws.cell(row=2, column=6).value is not None  # URL이 저장됨


def test_run_continues_after_single_record_failure(tmp_path: Path, monkeypatch: Any) -> None:
    monkeypatch.setenv("RUNPOD_API_KEY", "test-key")
    input_path = tmp_path / "students.xlsx"
    output_path = tmp_path / "students.updated.xlsx"
    _write_workbook(input_path, with_existing_url=False)

    service = FailingFirstCreateRunPodService()
    exit_code = run(
        _base_args(input_path, output_path),
        runpod_factory=lambda config, logger: service,
        jupyter_checker=lambda url, timeout: (True, "HTTP_200"),
    )

    # One row fails, but the next row is still processed and file is saved.
    assert exit_code == 1
    saved = load_workbook(output_path)
    ws = saved["시트1"]
    assert ws.cell(row=2, column=6).value is None
    assert ws.cell(row=3, column=6).value is not None  # URL이 저장됨


def test_run_uses_gpu_fallback_when_primary_unavailable(tmp_path: Path, monkeypatch: Any) -> None:
    monkeypatch.setenv("RUNPOD_API_KEY", "test-key")
    input_path = tmp_path / "students.xlsx"
    output_path = tmp_path / "students.updated.xlsx"
    _write_workbook(input_path, with_existing_url=False)

    args = _base_args(input_path, output_path)
    args.gpu_type = "NVIDIA A100 80GB PCIe"
    args.gpu_type_fallback = ["NVIDIA RTX 4090"]

    service = FallbackAwareRunPodService(unavailable_gpu="NVIDIA A100 80GB PCIe")
    exit_code = run(
        args,
        runpod_factory=lambda config, logger: service,
        jupyter_checker=lambda url, timeout: (True, "HTTP_200"),
    )

    assert exit_code == 0
    assert service.gpu_calls[0] == "NVIDIA A100 80GB PCIe"
    assert "NVIDIA RTX 4090" in service.gpu_calls


def test_run_uses_gpu_fallback_when_capacity_error_is_wrapped(tmp_path: Path, monkeypatch: Any) -> None:
    monkeypatch.setenv("RUNPOD_API_KEY", "test-key")
    input_path = tmp_path / "students.xlsx"
    output_path = tmp_path / "students.updated.xlsx"
    _write_workbook(input_path, with_existing_url=False)

    args = _base_args(input_path, output_path)
    args.gpu_type = "NVIDIA A100 80GB PCIe"
    args.gpu_type_fallback = ["NVIDIA RTX 4090"]

    service = WrappedFallbackAwareRunPodService(unavailable_gpu="NVIDIA A100 80GB PCIe")
    exit_code = run(
        args,
        runpod_factory=lambda config, logger: service,
        jupyter_checker=lambda url, timeout: (True, "HTTP_200"),
    )

    assert exit_code == 0
    assert service.gpu_calls[0] == "NVIDIA A100 80GB PCIe"
    assert "NVIDIA RTX 4090" in service.gpu_calls


def test_run_sync_updates_status_only(tmp_path: Path, monkeypatch: Any) -> None:
    monkeypatch.setenv("RUNPOD_API_KEY", "test-key")
    input_path = tmp_path / "students.xlsx"
    output_path = tmp_path / "students.updated.xlsx"
    _write_workbook(input_path, with_existing_url=False)

    args = _base_args(input_path, output_path)
    args.action = "sync"

    service = FakeRunPodService(existing_pods={
        "class-01-user01": {"id": "pod-existing", "name": "class-01-user01", "runtime": {"host": "pod.example.com", "port": 8888}}
    })
    exit_code = run(
        args,
        runpod_factory=lambda config, logger: service,
        jupyter_checker=lambda url, timeout: (True, "HTTP_200"),
    )

    assert exit_code == 0
    saved = load_workbook(output_path)
    ws = saved["시트1"]
    assert ws.cell(row=2, column=6).value is None
    assert ws.cell(row=2, column=7).value == "OK"


def test_run_sync_keeps_existing_url_when_pod_missing(tmp_path: Path, monkeypatch: Any) -> None:
    monkeypatch.setenv("RUNPOD_API_KEY", "test-key")
    input_path = tmp_path / "students.xlsx"
    output_path = tmp_path / "students.updated.xlsx"
    _write_workbook(input_path, with_existing_url=True)

    args = _base_args(input_path, output_path)
    args.action = "sync"

    service = FakeRunPodService(existing_pods={})
    exit_code = run(
        args,
        runpod_factory=lambda config, logger: service,
        jupyter_checker=lambda url, timeout: (True, "HTTP_200"),
    )

    assert exit_code == 0
    saved = load_workbook(output_path)
    ws = saved["시트1"]
    assert ws.cell(row=2, column=6).value == "https://existing-url"
    assert ws.cell(row=2, column=7).value == "POD_NOT_FOUND"


def test_run_lists_gpu_types_grouped_by_40gb(tmp_path: Path, monkeypatch: Any, capsys: Any) -> None:
    monkeypatch.setenv("RUNPOD_API_KEY", "test-key")
    input_path = tmp_path / "students.xlsx"
    output_path = tmp_path / "students.updated.xlsx"
    _write_workbook(input_path, with_existing_url=False)

    args = _base_args(input_path, output_path)
    args.list_gpu_types = True

    class ListingRunPodService(FakeRunPodService):
        def list_gpu_types_with_memory_and_secure_price(
            self,
            *,
            min_memory_gb: float | None = None,
        ) -> list[tuple[str, float, float | None]]:
            items = [
                ("NVIDIA A100 80GB PCIe", 80.0, 1.39),
                ("NVIDIA A40", 48.0, 0.79),
                ("NVIDIA RTX 4090", 24.0, 0.59),
                ("NVIDIA RTX 4000 Ada", 20.0, None),
            ]
            if min_memory_gb is None:
                return items
            return [item for item in items if item[1] >= min_memory_gb]

    service = ListingRunPodService(existing_pods={})

    exit_code = run(
        args,
        runpod_factory=lambda config, logger: service,
    )

    captured = capsys.readouterr()

    assert exit_code == 0
    assert "=== Memory >= 40GB ===" in captured.out
    assert "=== Memory < 40GB ===" in captured.out
    assert captured.out.index("NVIDIA A100 80GB PCIe | 80 GB") < captured.out.index("NVIDIA A40 | 48 GB")
    assert captured.out.index("NVIDIA RTX 4090 | 24 GB") < captured.out.index("NVIDIA RTX 4000 Ada | 20 GB")
    assert captured.out.index("=== Memory >= 40GB ===") < captured.out.index("=== Memory < 40GB ===")
    assert "NVIDIA RTX 4000 Ada | 20 GB | securePrice: 정보 없음" in captured.out


def test_run_stop_missing_pod_keeps_existing_url(tmp_path: Path, monkeypatch: Any) -> None:
    monkeypatch.setenv("RUNPOD_API_KEY", "test-key")
    input_path = tmp_path / "students.xlsx"
    output_path = tmp_path / "students.updated.xlsx"
    _write_workbook(input_path, with_existing_url=True)

    args = _base_args(input_path, output_path)
    args.action = "stop"
    args.numbers = [1]

    service = FakeRunPodService(existing_pods={})
    exit_code = run(
        args,
        runpod_factory=lambda config, logger: service,
        jupyter_checker=lambda url, timeout: (True, "HTTP_200"),
    )

    assert exit_code == 0
    saved = load_workbook(output_path)
    ws = saved["시트1"]
    assert ws.cell(row=2, column=6).value == "https://existing-url"
    assert ws.cell(row=2, column=7).value == "POD_NOT_FOUND"


def test_run_terminate_keeps_existing_url(tmp_path: Path, monkeypatch: Any) -> None:
    monkeypatch.setenv("RUNPOD_API_KEY", "test-key")
    input_path = tmp_path / "students.xlsx"
    output_path = tmp_path / "students.updated.xlsx"
    _write_workbook(input_path, with_existing_url=True)

    args = _base_args(input_path, output_path)
    args.action = "terminate"
    args.numbers = [1]

    service = FakeRunPodService(existing_pods={
        "class-01-user01": {"id": "pod-existing", "name": "class-01-user01", "runtime": {"host": "pod.example.com", "port": 8888}},
    })
    exit_code = run(
        args,
        runpod_factory=lambda config, logger: service,
        jupyter_checker=lambda url, timeout: (True, "HTTP_200"),
    )

    assert exit_code == 0
    saved = load_workbook(output_path)
    ws = saved["시트1"]
    assert ws.cell(row=2, column=6).value == "https://existing-url"
    assert ws.cell(row=2, column=7).value == "TERMINATED"


def test_run_provision_writes_gpu_type_column(tmp_path: Path, monkeypatch: Any) -> None:
    monkeypatch.setenv("RUNPOD_API_KEY", "test-key")
    input_path = tmp_path / "students.xlsx"
    output_path = tmp_path / "students.updated.xlsx"
    _write_workbook(input_path, with_existing_url=False)

    args = _base_args(input_path, output_path)
    args.gpu_type = "NVIDIA A100 80GB PCIe"

    service = FakeRunPodService(existing_pods={})
    exit_code = run(
        args,
        runpod_factory=lambda config, logger: service,
        jupyter_checker=lambda url, timeout: (True, "HTTP_200"),
    )

    assert exit_code == 0
    saved = load_workbook(output_path)
    ws = saved["시트1"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    gpu_col = headers.index("GPU Type") + 1
    assert ws.cell(row=2, column=gpu_col).value == "NVIDIA A100 80GB PCIe"


def test_run_provision_writes_ssh_column(tmp_path: Path, monkeypatch: Any) -> None:
    monkeypatch.setenv("RUNPOD_API_KEY", "test-key")
    input_path = tmp_path / "students.xlsx"
    output_path = tmp_path / "students.updated.xlsx"
    _write_workbook(input_path, with_existing_url=False)

    args = _base_args(input_path, output_path)
    service = FakeRunPodService(existing_pods={})

    exit_code = run(
        args,
        runpod_factory=lambda config, logger: service,
        jupyter_checker=lambda url, timeout: (True, "HTTP_200"),
    )

    assert exit_code == 0
    saved = load_workbook(output_path)
    ws = saved["시트1"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    ssh_col = headers.index("SSH") + 1
    assert ws.cell(row=2, column=ssh_col).value == "ssh generated-class-01-user01@ssh.runpod.io -i ~/.ssh/id_ed25519"


def test_run_stop_action_targets_specific_number(tmp_path: Path, monkeypatch: Any) -> None:
    monkeypatch.setenv("RUNPOD_API_KEY", "test-key")
    input_path = tmp_path / "students.xlsx"
    output_path = tmp_path / "students.updated.xlsx"
    _write_workbook(input_path, with_existing_url=False)

    args = _base_args(input_path, output_path)
    args.action = "stop"
    args.numbers = [1]

    service = FakeRunPodService(existing_pods={
        "class-01-user01": {"id": "pod-existing", "name": "class-01-user01", "runtime": {"host": "pod.example.com", "port": 8888}},
        "class-00-instructor": {"id": "pod-inst", "name": "class-00-instructor", "runtime": {"host": "pod.example.com", "port": 8888}},
    })

    exit_code = run(
        args,
        runpod_factory=lambda config, logger: service,
        jupyter_checker=lambda url, timeout: (True, "HTTP_200"),
    )

    assert exit_code == 0
    assert service.stopped_ids == ["pod-existing"]


def test_run_recreates_when_jupyter_unreachable(tmp_path: Path, monkeypatch: Any) -> None:
    monkeypatch.setenv("RUNPOD_API_KEY", "test-key")
    input_path = tmp_path / "students.xlsx"
    output_path = tmp_path / "students.updated.xlsx"
    _write_workbook(input_path, with_existing_url=False)

    args = _base_args(input_path, output_path)
    args.gpu_type_fallback = ["NVIDIA RTX 4090"]
    service = FakeRunPodService(existing_pods={})

    attempts = {"count": 0}

    def checker(url: str, timeout: int) -> tuple[bool, str]:
        attempts["count"] += 1
        if attempts["count"] == 1:
            return False, "HTTP_503"
        return True, "HTTP_200"

    args.jupyter_check = True
    args.recreate_on_unreachable = True
    exit_code = run(
        args,
        runpod_factory=lambda config, logger: service,
        jupyter_checker=checker,
    )

    assert exit_code == 0
    # first row does initial check + recreate check
    assert attempts["count"] >= 2
    user01_calls = [call["gpu_type_id"] for call in service.calls if call["name"] == "class-01-user01"]
    assert user01_calls[:2] == ["NVIDIA A100 80GB PCIe", "NVIDIA A100 80GB PCIe"]
    saved = load_workbook(output_path)
    ws = saved["시트1"]
    assert ws.cell(row=2, column=7).value == "UNREACHABLE_RECREATED"


def test_run_prefers_fallback_gpu_when_http_404_after_provision(tmp_path: Path, monkeypatch: Any) -> None:
    monkeypatch.setenv("RUNPOD_API_KEY", "test-key")
    input_path = tmp_path / "students.xlsx"
    output_path = tmp_path / "students.updated.xlsx"
    _write_workbook(input_path, with_existing_url=False)

    args = _base_args(input_path, output_path)
    args.gpu_type = "NVIDIA A100 80GB PCIe"
    args.gpu_type_fallback = ["NVIDIA A100-SXM4-80GB", "NVIDIA RTX 4090"]
    args.jupyter_check = True
    args.recreate_on_unreachable = True

    service = FakeRunPodService(existing_pods={})
    attempts = {"count": 0}

    def checker(url: str, timeout: int) -> tuple[bool, str]:
        attempts["count"] += 1
        if attempts["count"] == 1:
            return False, "HTTP_404"
        return True, "HTTP_200"

    exit_code = run(
        args,
        runpod_factory=lambda config, logger: service,
        jupyter_checker=checker,
    )

    assert exit_code == 0
    user01_calls = [call["gpu_type_id"] for call in service.calls if call["name"] == "class-01-user01"]
    assert user01_calls[:2] == ["NVIDIA A100 80GB PCIe", "NVIDIA A100-SXM4-80GB"]
    saved = load_workbook(output_path)
    ws = saved["시트1"]
    assert ws.cell(row=2, column=7).value == "UNREACHABLE_RECREATED"


def test_run_marks_unreachable_without_recreate_when_disabled(tmp_path: Path, monkeypatch: Any) -> None:
    monkeypatch.setenv("RUNPOD_API_KEY", "test-key")
    input_path = tmp_path / "students.xlsx"
    output_path = tmp_path / "students.updated.xlsx"
    _write_workbook(input_path, with_existing_url=False)

    args = _base_args(input_path, output_path)
    args.jupyter_check = True  # 접속 체크 활성화, 재생성은 기본값(False) 유지

    service = FakeRunPodService(existing_pods={})
    exit_code = run(
        args,
        runpod_factory=lambda config, logger: service,
        jupyter_checker=lambda url, timeout: (False, "HTTP_503"),
    )

    assert exit_code == 0
    saved = load_workbook(output_path)
    ws = saved["시트1"]
    assert ws.cell(row=2, column=7).value == "UNREACHABLE"


def test_run_skip_jupyter_check_is_fast_and_marks_skipped(tmp_path: Path, monkeypatch: Any) -> None:
    monkeypatch.setenv("RUNPOD_API_KEY", "test-key")
    input_path = tmp_path / "students.xlsx"
    output_path = tmp_path / "students.updated.xlsx"
    _write_workbook(input_path, with_existing_url=False)

    args = _base_args(input_path, output_path)
    # skip_jupyter_check=True 가 기본값이므로 별도 플래그 불필요
    service = FakeRunPodService(existing_pods={})

    def should_not_be_called(url: str, timeout: int) -> tuple[bool, str]:
        raise AssertionError("jupyter checker should not be called when skip_jupyter_check=True")

    exit_code = run(
        args,
        runpod_factory=lambda config, logger: service,
        jupyter_checker=should_not_be_called,
    )

    assert exit_code == 0
    saved = load_workbook(output_path)
    ws = saved["시트1"]
    assert ws.cell(row=2, column=7).value == "CHECK_SKIPPED"



